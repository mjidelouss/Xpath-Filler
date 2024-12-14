import openpyxl
from lxml import etree
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

class XPathCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("XPath Checker")
        self.root.geometry("600x400")
        
        # Create main frame
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Input file selection
        ttk.Label(main_frame, text="Input Excel File:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.input_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.input_path, width=50).grid(row=0, column=1, padx=5)
        ttk.Button(main_frame, text="Browse", command=self.browse_input).grid(row=0, column=2)
        
        # XML folder selection
        ttk.Label(main_frame, text="XML Folder:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.xml_folder_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.xml_folder_path, width=50).grid(row=1, column=1, padx=5)
        ttk.Button(main_frame, text="Browse", command=self.browse_xml_folder).grid(row=1, column=2)
        
        # Output file selection
        ttk.Label(main_frame, text="Output Excel File:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.output_path = tk.StringVar()
        ttk.Entry(main_frame, textvariable=self.output_path, width=50).grid(row=2, column=1, padx=5)
        ttk.Button(main_frame, text="Browse", command=self.browse_output).grid(row=2, column=2)
        
        # Progress bar
        self.progress = ttk.Progressbar(main_frame, length=400, mode='indeterminate')
        self.progress.grid(row=3, column=0, columnspan=3, pady=20)
        
        # Status label
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = ttk.Label(main_frame, textvariable=self.status_var)
        self.status_label.grid(row=4, column=0, columnspan=3)
        
        # Process button
        self.process_button = ttk.Button(main_frame, text="Process", command=self.start_processing)
        self.process_button.grid(row=5, column=0, columnspan=3, pady=20)

    def browse_input(self):
        filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filename:
            self.input_path.set(filename)
            # Auto-generate output filename
            output = filename.rsplit('.', 1)[0] + '_output.xlsx'
            self.output_path.set(output)

    def browse_xml_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.xml_folder_path.set(folder)

    def browse_output(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if filename:
            self.output_path.set(filename)

    def start_processing(self):
        if not self.validate_inputs():
            return
        
        self.process_button.state(['disabled'])
        self.progress.start()
        self.status_var.set("Processing...")
        
        # Start processing in a separate thread
        thread = threading.Thread(target=self.process_files)
        thread.daemon = True
        thread.start()

    def validate_inputs(self):
        if not self.input_path.get():
            messagebox.showerror("Error", "Please select an input Excel file")
            return False
        if not self.xml_folder_path.get():
            messagebox.showerror("Error", "Please select an XML folder")
            return False
        if not self.output_path.get():
            messagebox.showerror("Error", "Please specify an output Excel file")
            return False
        return True

    def process_files(self):
        try:
            xml_files = [os.path.join(self.xml_folder_path.get(), f) 
                        for f in os.listdir(self.xml_folder_path.get()) 
                        if f.endswith('.xml')]
            
            results, original_workbook = process_data(self.input_path.get(), xml_files)
            write_excel(results, original_workbook, self.output_path.get())
            
            self.root.after(0, self.processing_complete)
        except Exception as e:
            self.root.after(0, lambda: self.processing_error(str(e)))

    def processing_complete(self):
        self.progress.stop()
        self.process_button.state(['!disabled'])
        self.status_var.set("Processing complete!")
        messagebox.showinfo("Success", f"Results written to {self.output_path.get()}")

    def processing_error(self, error_message):
        self.progress.stop()
        self.process_button.state(['!disabled'])
        self.status_var.set("Error occurred!")
        messagebox.showerror("Error", f"An error occurred: {error_message}")

def main():
    root = tk.Tk()
    app = XPathCheckerApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()

# Keep the existing functions unchanged
def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data, workbook

def check_xpath(xpath, xml_files):
    for xml_file in xml_files:
        try:
            tree = etree.parse(xml_file)
            result = tree.xpath(xpath)
            if result:
                return True
        except:
            pass
    return False

def process_data(input_file, xml_files):
    data, workbook = read_excel(input_file)
    headers = data[0]
    xpath_column = headers.index("XPath")  # Assuming "XPath" is the column name

    # Check if result columns already exist
    new_columns = ["Donnees manquant", "Action Cegedim", "Action DO"]
    existing_columns = [col for col in new_columns if col in headers]
    
    for col in new_columns:
        if col not in headers:
            headers.append(col)
    
    results = [headers]

    for row in data[1:]:
        xpath = row[xpath_column]
        xpath_exists = check_xpath(xpath, xml_files)
        
        new_row = list(row)
        if xpath_exists:
            new_values = ["NON", "Aucun action", "Fournir la donnee"]
        else:
            new_values = ["OUI", "Fournir la donnee", "Aucun action"]
        
        for i, col in enumerate(new_columns):
            if col in existing_columns:
                index = headers.index(col)
                new_row[index] = new_values[i]
            else:
                new_row.append(new_values[i])
        
        results.append(new_row)

    return results, workbook

def write_excel(data, original_workbook, output_file):
    new_workbook = openpyxl.Workbook()
    sheet = new_workbook.active
    for row in data:
        sheet.append(row)
    
    # Copy styles from original workbook
    for i, row in enumerate(original_workbook.active.iter_rows(min_row=1, max_row=len(data), max_col=len(data[0])-3), start=1):
        for j, cell in enumerate(row, start=1):
            new_cell = sheet.cell(row=i, column=j)
            new_cell.font = cell.font.copy()
            new_cell.border = cell.border.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection.copy()
            new_cell.alignment = cell.alignment.copy()

    new_workbook.save(output_file)